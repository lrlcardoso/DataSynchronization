"""
==============================================================================
Title:          Signal Synchronization Pipeline (Folder-Based)
Description:    Aligns and synchronizes IMU and video marker data for all
                segment folders found inside the input directory. Selects
                optimal camera per segment and saves aligned CSVs and plots.
Author:         Lucas R. L. Cardoso
Project:        VRRehab_UQ-MyTurn
Date:           2025-04-25
Version:        1.1
==============================================================================
Usage:
    python main_sync.py

Dependencies:
    - Python >= 3.x
    - Required libraries: os, time, pandas, openpyxl, natsort
    - Local modules: config, utils (file_utils, signal_processing, plotting)

Notes:
    - The script processes all selected patients/sessions/segments defined in 
      `config.py`.
    - Each segment is processed independently, with optimal camera selected 
      based on marker visibility.
    - Results include lag/correlation values and resynchronized CSVs and plots.

Changelog:
    - v1.0: [2025-04-25] Initial implementation of folder-based synchronization.
    - v1.1: [2025-07-27] Added support for arbitrary patient/session/camera/segment
                         combinations. Also added saving of output CSVs and plots
                         for both cameras and all IMU loggers.
==============================================================================
"""

import os
import time
import openpyxl
import pandas as pd
from natsort import natsorted
from openpyxl import load_workbook, Workbook
from collections import defaultdict

sync_results = defaultdict(lambda: defaultdict(dict))  # {session: {segment: {"lag": ..., "corr": ...}}}

# === Project Modules ===
from config import (
    ROOT_DIR,
    SELECTED_PATIENTS,
    SELECTED_SESSIONS,
    SELECTED_SUBFOLDERS,
    SELECTED_SEGMENTS,
    IMU_FREQ,
    VIDEO_FREQ,
    FILTER_LOW_CUT,
    FILTER_HIGH_CUT,
    FILTER_ORDER,
    LAG_RANGE,
    WINDOW_SIZE,
    SHOW_PLOTS,
    SHOW_DEBG_PLOTS,
    OUTPUT_DIR,
    SAVE_PLOTS,
    SAVE_CSV
)
from utils.file_utils import (
    load_imu_data, 
    get_videos_info
)
from utils.signal_processing import (
    smooth_signal, 
    resample_signal,
    compute_magnitude, 
    highpass_filter, 
    normalize_signal,
    lowpass_filter, 
    position_to_acceleration,
    compute_cross_correlation,
    align_signals,apply_lag
)
from utils.plotting import (
    plot_debug, 
    plot_and_save_similarity,
    plot_and_save_magnitude
)

def run_sync(video_path, patient, session, affected_side):

    print("="*100)
    print(f"📝 Processing: {patient} ({affected_side}) | {session}")
    print("="*100)

    seg_dir = os.path.join(video_path, "Camera1", "Segments")

    all_seg_names = natsorted(os.listdir(seg_dir))

    if SELECTED_SEGMENTS is None:
        seg_names = all_seg_names
    else:
        seg_names = [all_seg_names[i] for i in SELECTED_SEGMENTS if i < len(all_seg_names)]

    for seg_name in seg_names:
        if "static" in seg_name.lower():
            print(f"Skipped segment {seg_name}.")
            continue  # Skip 'static' segments

        # Find in which camera we have more markers visible
        if affected_side == "R" or affected_side == "B":
            marker = "11_"
        elif affected_side == "L":
            marker = "10_"
        else:
            raise ValueError(f"Unknown affected_side: {affected_side}")
        
        best_camera = None
        max_visible_points = -1
        best_df = None

        for camera_name in ["Camera1", "Camera2"]:
            camera_path = os.path.join(video_path, camera_name, "Segments", seg_name)
            if not os.path.isdir(camera_path):
                continue  # Skip if camera folder is missing

            csv_files = [f for f in os.listdir(camera_path) if f.endswith('.csv')]
            if not csv_files:
                raise FileNotFoundError(f"No CSV file found in folder: {camera_path}")
            if len(csv_files) > 1:
                raise RuntimeError(f"Multiple CSV files found in folder: {camera_path} → {csv_files}")

            csv_path = os.path.join(camera_path, csv_files[0])
            df = pd.read_csv(csv_path)
            if df.empty:
                raise ValueError(f"CSV file is empty: {csv_path}")

            x_col, y_col = f"{marker}x", f"{marker}y"
            if x_col not in df.columns or y_col not in df.columns:
                continue  # Skip if marker columns are missing

            visible_count = ((df[x_col] != 0) & (df[y_col] != 0)).sum()

            if visible_count > max_visible_points:
                max_visible_points = visible_count
                best_camera = camera_name
                best_df = df

        if best_camera is None or best_df is None:
            raise ValueError(f"No valid camera data found for segment: {seg_name}")
    
        print()
        print(f"📂 Segment {seg_name} ({best_camera})")
        print("-" * 100)

        # 0 - Create the paths to save later
        plots_dir = os.path.join(
            ROOT_DIR,
            OUTPUT_DIR,
            patient,
            session,
            seg_name,
            "Plots"
        )
        os.makedirs(plots_dir, exist_ok=True)

        csvs_dir = os.path.join(
            ROOT_DIR,
            OUTPUT_DIR,
            patient,
            session,
            seg_name,
        )
        os.makedirs(csvs_dir, exist_ok=True)

        # 1 - Load all necessary data and info
        # Video data
        video_data = best_df[['Unix Time', f"{marker}x", f"{marker}y"]].copy()
        video_data['Unix Time'] = video_data['Unix Time']
        video_data[f"{marker}y"] = -video_data[f"{marker}y"]
        # IMU data
        # Get UNIX start and end from first/last row
        seg_start = video_data.iloc[0]['Unix Time']
        seg_end = video_data.iloc[-1]['Unix Time']
        seg_interval = (seg_start, seg_end)
        imu_path = os.path.abspath(os.path.join(os.path.dirname(video_path), "..", "WMORE"))
        imu_data = load_imu_data(imu_path, affected_side, seg_interval)
        imu_data['ax'] = -imu_data['ax']

        # 2 - Preprocessing
        # Resample video_data to make sure there is no gaps. Also add 0 to every gap to facilitate visualization
        video_data_resampled = resample_signal(video_data, target_freq=VIDEO_FREQ, fill_missing_with_nan=True)
        
        # Apply a lowpass filter to reduce noise, but do it avoiding the chunks in which the signal is zero (bad detection)
        video_data_lowpass_filtered = lowpass_filter(video_data_resampled, fs=VIDEO_FREQ, cutoff=FILTER_HIGH_CUT, order=FILTER_ORDER)

        if SHOW_DEBG_PLOTS:
            plot_debug(video_data_resampled, video_data_lowpass_filtered, markers=(f"{marker}y",f"{marker}y"), labels=["Video - resampled", "Video - filtered"])

        # Apply a highpass filter to match to the IMU that must be highpassed filtered to get rid of the gravity effect
        video_data_bandpass_filtered = highpass_filter(video_data_lowpass_filtered, fs=VIDEO_FREQ, cutoff=FILTER_LOW_CUT, order=FILTER_ORDER)

        # Resample to IMU data to make sure there is no gap
        imu_data_resampled = resample_signal(imu_data, target_freq=IMU_FREQ, fill_missing_with_nan=True)

        # Apply a lowpass filter to reduce noise, matching the video_data
        imu_data_lowpass_filtered = lowpass_filter(imu_data_resampled, fs=IMU_FREQ, cutoff=FILTER_HIGH_CUT, order=FILTER_ORDER)

        if SHOW_DEBG_PLOTS:
            plot_debug(imu_data_resampled, imu_data_lowpass_filtered, markers=("ax","ax"), labels=["IMU - resampled", "IMU - filtered"])

        # Apply a highpass filter to remove gravity
        imu_data_bandpass_filtered = highpass_filter(imu_data_lowpass_filtered, fs=IMU_FREQ, cutoff=FILTER_HIGH_CUT, order=FILTER_ORDER)

        # 3 - Compute magnitudes
        video_mag = position_to_acceleration(video_data_bandpass_filtered, f"{marker}x", f"{marker}y")
        imu_mag = compute_magnitude(imu_data_bandpass_filtered)

        # 4 - Downsample IMU data to match video data
        imu_rs = align_signals(video_mag, imu_mag, method='interp')

        # 5 - Smooth signals
        video_smooth = smooth_signal(video_mag, window=WINDOW_SIZE)
        imu_smooth = smooth_signal(imu_rs, window=WINDOW_SIZE)
        # video_smooth = video_mag
        # imu_smooth = imu_rs
        
        if SHOW_DEBG_PLOTS:
            plot_debug(video_mag, video_smooth, markers=("Magnitude","Magnitude"), labels=["Video - magnitude", "Video - smoothed"])
            plot_debug(imu_mag, imu_smooth, markers=("Magnitude","Magnitude"), labels=["IMU - magnitude", "IMU - smoothed"])

        # 6 - Normalize
        video_norm = normalize_signal(video_smooth)
        imu_norm = normalize_signal(imu_smooth)

        if SHOW_DEBG_PLOTS:
            plot_debug(video_norm, imu_norm,  markers=("Magnitude","Magnitude") , labels=["Video", "IMU"])

        # 7 - Cross-correlate
        lag_samples, max_corr, similarity_curve = compute_cross_correlation(imu_norm, video_norm, VIDEO_FREQ, LAG_RANGE)
        print(f"Optimal lag (s): {lag_samples/VIDEO_FREQ:.3f}, Correlation: {max_corr:.3f}")
        
        # Save to in-memory results dict
        sync_results[session][seg_name] = {
            "lag": round(lag_samples / VIDEO_FREQ, 3),
            "corr": round(max_corr, 3)
        }

        # save plot similarity_curve
        plot_and_save_similarity(similarity_curve, LAG_RANGE, best_camera, output_dir=plots_dir, save=SAVE_PLOTS, show=SHOW_PLOTS)
        
        # 8 - Apply the lag and save CSV for both cameras
        global_start = None
        global_end = None
        for camera_name in ["Camera1", "Camera2"]:
            camera_path = os.path.join(video_path, camera_name, "Segments", seg_name)
            if not os.path.isdir(camera_path):
                print(f"⚠️  Skipping: {camera_name} folder not found for {patient}, {session}.")
                continue  # Skip if camera folder is missing

            csv_files = [f for f in os.listdir(camera_path) if f.endswith('.csv')]
            if not csv_files:
                raise FileNotFoundError(f"No CSV file found in folder: {camera_path}")
            if len(csv_files) > 1:
                raise RuntimeError(f"Multiple CSV files found in folder: {camera_path} → {csv_files}")

            csv_path = os.path.join(camera_path, csv_files[0])
            df = pd.read_csv(csv_path)
            if df.empty:
                raise ValueError(f"CSV file is empty: {csv_path}")

            data = resample_signal(df, target_freq=VIDEO_FREQ, fill_missing_with_nan=True)
            video_data_lagged = apply_lag(data, lag_samples, VIDEO_FREQ)

            # Track earliest start and latest end
            cam_start = video_data_lagged["Unix Time"].iloc[0]
            cam_end = video_data_lagged["Unix Time"].iloc[-1]

            global_start = cam_start if global_start is None else min(global_start, cam_start)
            global_end = cam_end if global_end is None else max(global_end, cam_end)

            if SAVE_CSV:
                filename = f"{camera_name}.csv"
                outpath = os.path.join(csvs_dir, filename)
                os.makedirs(csvs_dir, exist_ok=True)
                video_data_lagged.to_csv(outpath, index=False)
                print(f"✅ Saved {camera_name}.csv")

        # 9 - Save CSV for each logger, segment wise
        for i in range(1, 6):  # Logger1 to Logger5
            logger = f"Logger{i}"
            file_path = os.path.join(imu_path, f"{logger}.csv")

            if not os.path.exists(file_path):
                print(f"⚠️  Skipping: {logger} not found.")
                continue

            # Read the IMU CSV
            df = pd.read_csv(file_path)

            # Filter by segment interval
            global_start, global_end = seg_interval
            data_segmented = df[(df['Unix Time'] >= global_start) & (df['Unix Time'] <= global_end)].reset_index(drop=True)

            data_segmented_resampled = resample_signal(data_segmented, target_freq=IMU_FREQ, fill_missing_with_nan=True)

            # Save if needed
            if SAVE_CSV:
                filename = f"{logger}.csv"
                outpath = os.path.join(csvs_dir, filename)
                os.makedirs(csvs_dir, exist_ok=True)
                data_segmented_resampled.to_csv(outpath, index=False)
                print(f"✅ Saved {logger}.csv")

        # 10 - Plot for visual inspection
        # First determine which logger will be saved for later plotting
        if affected_side == "R" or affected_side == "B":
            logger = "Logger1"
        elif affected_side == "L":
            logger = "Logger2"

        # Load IMU data
        file_path = os.path.join(csvs_dir, f"{logger}.csv")

        if not os.path.exists(file_path):
            print(f"⚠️  Warning: CSV file not found: {file_path}")
            continue  # Skip to next camera

        imu_data_saved = pd.read_csv(file_path)

        if imu_data_saved.empty:
            raise ValueError(f"CSV file is empty: {csv_path}")

        # Load video data
        for camera_name in ["Camera1", "Camera2"]:
            file_path = os.path.join(csvs_dir, f"{camera_name}.csv")

            if not os.path.exists(file_path):
                print(f"⚠️  Warning: CSV file not found: {file_path}")
                continue  # Skip to next camera

            video_data_saved = pd.read_csv(file_path)

            if video_data_saved.empty:
                raise ValueError(f"CSV file is empty: {csv_path}")

            # Regenerate magnitudes with the saved data, and plot/save graphs for visual inspection 
            video_data_saved_lowpass_filtered = lowpass_filter(video_data_saved, fs=VIDEO_FREQ, cutoff=FILTER_HIGH_CUT, order=FILTER_ORDER)
            imu_data_saved_lowpass_filtered = lowpass_filter(imu_data_saved, fs=IMU_FREQ, cutoff=FILTER_HIGH_CUT, order=FILTER_ORDER)
            
            video_data_saved_bandpass_filtered = highpass_filter(video_data_saved_lowpass_filtered, fs=VIDEO_FREQ, cutoff=FILTER_LOW_CUT, order=FILTER_ORDER)
            imu_data_saved_bandpass_filtered = highpass_filter(imu_data_saved_lowpass_filtered, fs=IMU_FREQ, cutoff=FILTER_HIGH_CUT, order=FILTER_ORDER)

            video_data_saved_mag = position_to_acceleration(video_data_saved_bandpass_filtered, f"{marker}x", f"{marker}y")
            imu_data_saved_mag = compute_magnitude(imu_data_saved_bandpass_filtered)

            imu_data_saved_rs = align_signals(video_data_saved_mag, imu_data_saved_mag, method='interp')

            video_data_saved_smooth = smooth_signal(video_data_saved_mag, window=WINDOW_SIZE)
            imu_data_saved_smooth = smooth_signal(imu_data_saved_rs, window=WINDOW_SIZE)
            # video_data_saved_smooth = video_data_saved_mag
            # imu_data_saved_smooth = imu_data_saved_rs
            
            video_data_norm = normalize_signal(video_data_saved_smooth)
            imu_data_saved_norm = normalize_signal(imu_data_saved_smooth)

            #save plots video_norm vs imu_norm
            plot_and_save_magnitude(video_data_norm, imu_data_saved_norm, camera_name, output_dir=plots_dir, save=SAVE_PLOTS, show=SHOW_PLOTS)

def main():
    """
    Entry point for batch video processing.
    Select the segments to run.
    """
    start_time_all = time.time()

    video_info = get_videos_info(ROOT_DIR, SELECTED_PATIENTS, SELECTED_SESSIONS, SELECTED_SUBFOLDERS)
    
    for video_path, patient, session, affected_side in video_info:
        run_sync(video_path, patient, session, affected_side)

    print("-" * 100)
    print(f"All segments were sync in {time.strftime('%H:%M:%S', time.gmtime(time.time() - start_time_all))}.")
    print("-" * 100)

    excel_out_path = os.path.join("SyncSummary.xlsx")

    # Load existing workbook if it exists, otherwise create new
    if os.path.exists(excel_out_path):
        wb = load_workbook(excel_out_path)
    else:
        wb = Workbook()

    SESSION_COLUMN_MAP = {
        "Session1": 1,  # A
        "Session2": 4,  # D
        "Session3": 7,  # G
    }

    for patient in SELECTED_PATIENTS:
        if patient in wb.sheetnames:
            ws = wb[patient]
        else:
            ws = wb.create_sheet(title=patient)

        for session in sync_results:
            session_prefix = session.split("_")[0]  # e.g., "Session1" from "Session1_20250305"

            if session_prefix not in SESSION_COLUMN_MAP:
                print(f"⚠️ Unknown session prefix '{session_prefix}' — skipping.")
                continue

            col_base = SESSION_COLUMN_MAP[session_prefix]

            ws.cell(row=2, column=col_base, value=session)
            ws.cell(row=3, column=col_base, value="Segment")
            ws.cell(row=3, column=col_base + 1, value="Lag")
            ws.cell(row=3, column=col_base + 2, value="Correlation")

            row = 4
            for segment, values in sync_results[session].items():
                ws.cell(row=row, column=col_base, value=segment)
                ws.cell(row=row, column=col_base + 1, value=values["lag"])
                ws.cell(row=row, column=col_base + 2, value=values["corr"])
                row += 1

    # Remove default sheet if empty
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    wb.save(excel_out_path)
    print(f"Saved Excel summary to {excel_out_path}")

if __name__ == "__main__":
    main()
